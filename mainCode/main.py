# -*- coding: utf-8 -*-
# python39
from WeChatPYAPI import WeChatPYApi

import json, requests, win32con, win32api, psutil
from Cryptodome.Cipher import AES
try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

import threading

import time
import os
import re
import sys

import logging
from queue import Queue
import openpyxl
from pathlib import Path
import webbrowser
import pandas as pd
import requests

from tkinter import *
from tkinter import filedialog
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
from tkinter import scrolledtext

import math
import numpy as np
# 注意要放在tk后import 不然出错
try:
    from PIL import Image
except ImportError:
    import Image

import sys

'''
if hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")
# print(base_path)

application_path = ""
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)  # 打包EXE后的运行路径
elif __file__:
    application_path = os.path.dirname(__file__)  # 未打包时的运行路径
# print(application_path)
'''

# 创建线程执行程序，防止功能函数阻塞
def thread_it(func, *args):		# 传入函数名和参数
    # 创建线程
    t = threading.Thread(target=func, args=args)
    # 守护线程
    t.setDaemon(True)
    # 启动
    t.start()

#  获取GUI文件路径
ASSETS_PATH = Path(__file__).resolve().parent / "assets"
def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)

# 获取当前目录路径
BASE_DIR = os.getcwd()
print("BASE_DIR:",BASE_DIR)

# 获取电脑文件
path = getattr(sys, '_MEIPASS', os.getcwd())
os.chdir(path)

# 日志器
logging.basicConfig(level=logging.INFO)

# 消息队列 
msg_queue = Queue()  # 消息队列

# 消息回调，建议异步处理，防止阻塞
def on_message(msg):
    # print(msg)
    msg_queue.put(msg)

# 退出事件回调
def on_exit(wx_id):
    print("已退出：{}".format(wx_id))

# 根据端口号杀死进程
def kill_port_process(port):
    ret = os.popen("netstat -nao|findstr " + str(port))
    str_list = ret.read()
    if not str_list:
        print('端口未使用')
        return
    # 只关闭处于LISTENING的端口
    if 'TCP' in str_list:
        ret_list = str_list.replace(' ', '')
        ret_list = re.split('\n', ret_list)
        listening_list = [rl.split('LISTENING') for rl in ret_list]
        process_pids = [ll[1] for ll in listening_list if len(ll) >= 2]
        process_pid_set = set(process_pids)
        for process_pid in process_pid_set:
            os.popen('taskkill /pid ' + str(process_pid) + ' /F')
            print(port, '端口已被释放')
            time.sleep(1)

    elif 'UDP' in str_list:
        ret_list = re.split(' ', str_list)
        process_pid = ret_list[-1].strip()
        if process_pid:
            os.popen('taskkill /pid ' + str(process_pid) + ' /F')
            print('端口已被释放')
        else:
            print("端口未被使用")

# ***********************初始化GUI，构建日志*****************************

# 初始化GUI
window = Tk()

logo = PhotoImage(file=relative_to_assets("logo.png"))
window.call('wm', 'iconphoto', window._w, logo)
window.title("微信好友助手")

window.geometry("862x519")
window.configure(bg="#F7F7F7")

canvas = Canvas(
    window,
    bg="#F7F7F7",
    height=519,
    width=862,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)

canvas.place(x=0, y=0)


# 创建一个滚动条控件，默认为垂直方向
sbar1= Scrollbar(window)
# 将滚动条放置在右侧，并设置当窗口大小改变时滚动条会沿着垂直方向延展
sbar1.pack(side=RIGHT, fill=Y)

# 构建日志输入框
entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
print(relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    629.5000000000001,
    354.4999999999999,
    image=entry_image_3
)
entry_3 = Text(
    bd=0,
    bg="#F2F2F2",
    fg="#000716",
    highlightthickness=0,
    yscrollcommand = sbar1.set
)
sbar1.config(command=entry_3.yview)

entry_3.place(
    x=436.0000000000001,
    y=216.9999999999999,
    width=387.0,
    height=273.0
)

# 保存日志到日志框，用entry组件
def notice(text):
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # 保存操作时间 #
    entry_3.configure(state='normal')
    entry_3.insert('insert', '[' + now + ']' + ':' + text + '\n' + '\n')  # 带上操作时间 #
    entry_3.update()
    entry_3.configure(state='disabled')
    entry_3.see(END)

# ***********************初始化API，进行登陆*****************************

# 构建API实例
def APIbegin():
    w = WeChatPYApi(msg_callback=on_message, exit_callback=on_exit, logger=logging)
    return w

# 每次启动前都杀死8888端口
kill_port_process(8888)

# 实例化WechatAPI对象,
w = APIbegin()

# 全局变量,初始化微信登录
self_wx = ""
my_info = ""
self_avatar_link = ""
loginFlag = 0

# 因为要等待5s，登录前在日志中进行提示
def printLoginBegin():
    notice("正在启动微信中，请稍候")
    time.sleep(0.5)

# 登录函数
def loginWechat(w):
    global loginFlag

    # 如果loginFlag==1，说明已经登录了，直接退出，否则继续
    if loginFlag:
        notice('您已经登录了微信，无需重复登录')
        return

    # 启动微信
    w.start_wx()

    # 网页端构建可以选择保存登录二维码
    # w.start_wx(path=os.path.join(BASE_DIR, "login_qrcode.png"))  

    # 这里需要阻塞，等待获取个人信息
    while not w.get_self_info():
        time.sleep(5)

    # 或者个人信息和wechat ID
    my_info = w.get_self_info()
    self_wx = my_info["wx_id"]
    self_avatar_link = my_info["avatar_url"]

    # 日志提示
    if my_info and self_wx:
        notice("微信登录成功")
        notice("登录信息为：%s" % (my_info))
        notice("---------------------")
        loginFlag = 1
    else:
        notice("出错了，请重新进行登录")


# 绑定到tk的button_1
button_image_1 = PhotoImage(
    file=relative_to_assets("button_1.png"))

button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    # command=lambda: print("button_1 clicked"),
    command=lambda: [printLoginBegin(), thread_it(loginWechat,w)],
    relief="flat"
)
button_1.place(
    x=71.00000000000011,
    y=107.99999999999989,
    width=120.0,
    height=33.0
)


# ***********************导出好友列表*****************************

# 将好友列表导出为 friends_list.xlsx
def list2excel(your_list):
    # Open a new workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    # Add the header row
    worksheet.append(['wx_id', 'nick_name', 'remark', 'wx_account', 'avatar_url'])
    # Add the data rows
    for item in your_list:
        worksheet.append([item['wx_id'], item['nick_name'], item['remark'], item['wx_account'], item['avatar_url']])
    # Save the workbook
    workbook.save('friends_list.xlsx')

# 因为要等待5s，登录前在日志中进行提示
def printFriendBegin():
    notice("正在导出好友列表中，请稍候")
    time.sleep(0.5)

# 导出好友列表
def friendList(w):
    if loginFlag == 1:
        lists = w.pull_list(self_wx=self_wx, pull_type=1)
        list2excel(lists)
        notice("好友列表导出完成，文件路径为：%s\\friends_list.xlsx" % (BASE_DIR))
        notice("---------------------")
    else:
        notice("您仍未登录，请先登录微信")


# 将导出好友列表函数绑定到button_2
button_image_2 = PhotoImage(
    file=relative_to_assets("button_2.png"))
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    # command=lambda: print("button_2 clicked"),
    command=lambda: [printFriendBegin(), thread_it(friendList,w)],
    relief="flat"
)
button_2.place(
    x=223.0000000000001,
    y=107.99999999999989,
    width=120.0,
    height=33.0
)

# ***********************上传excel文件和图片文件*****************************

# 全局变量 excel_path 和 image_path 存储供后续发送使用
excel_path = ""
image_path = ""

# 选择excel文件（button_5 以及 entery_1）
def select_excel():
    global excel_path

    # 打开文件选择对话框
    excel_path = filedialog.askopenfilename()
    # 清空控件
    entry_1.delete(0, END)
    # 将文件路径显示在输入框中
    entry_1.insert(0, excel_path)


# 表格路径展示
entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    250.0000000000001,
    228.9999999999999,
    image=entry_image_1
)
entry_1 = Entry(
    bd=0,
    bg="#F2F2F2",
    fg="#000716",
    highlightthickness=0
)
entry_1.place(
    x=130.0000000000001,
    y=216.9999999999999,
    width=240.0,
    height=22.0
)

# excel路径
excel_path = entry_1.get()
excel_path = excel_path.strip()

# 存储要发送的信息模板
data = []

# 读取编辑好将要发送信息的excel
def read_data_from_excel(filename):
    # Open the workbook and get the sheet
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    # Iterate over the rows of the sheet and build the dictionaries
    for row in worksheet.iter_rows(min_row=2):  # Skip the first row (header row)
        item = {}
        for cell in row:
            item[cell.column] = cell.value
        data.append(item)
    return data


def showExcel():
    global data
    if excel_path != "":
        data = read_data_from_excel(excel_path)
        print(data)
        notice(f'Excel文件读取成功，路径为：{excel_path} ')
        notice("请确认下面是否是您要发送的信息")
        notice("----------beigin--------------")
        notice(str(data))
        notice("---------- end--------------")
    else:
        notice("仍未选择excel文件，请您选择要发送的信息文件")

# 上传excel按钮
button_image_5 = PhotoImage(
    file=relative_to_assets("button_5.png"))
button_5 = Button(
    image=button_image_5,
    borderwidth=0,
    highlightthickness=0,
    # command=lambda: print("button_5 clicked"),
    command=lambda: [select_excel(), thread_it(showExcel)],
    relief="flat"
)
button_5.place(
    x=31.000000000000114,
    y=216.9999999999999,
    width=90.0,
    height=24.0
)


# 选择图片文件（button_6 以及 entery_2）
def select_image():
    global image_path

    # 打开文件选择对话框
    image_path = filedialog.askopenfilename()
    # 清空控件
    entry_2.delete(0, END)
    # 将文件路径显示在输入框中
    entry_2.insert(0, image_path)
    notice(f'图片读取成功，路径为：{image_path} ')


# 图片路径展示
entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    250.0000000000001,
    267.9999999999999,
    image=entry_image_2
)
entry_2 = Entry(
    bd=0,
    bg="#F2F2F2",
    fg="#000716",
    highlightthickness=0
)
entry_2.place(
    x=130.0000000000001,
    y=255.9999999999999,
    width=240.0,
    height=22.0
)
# image路径
image_path = entry_2.get()
image_path = image_path.strip()

# 图片上传按钮
button_image_6 = PhotoImage(
    file=relative_to_assets("button_6.png"))
button_6 = Button(
    image=button_image_6,
    borderwidth=0,
    highlightthickness=0,
    # command=lambda: print("button_6 clicked"),
    command=lambda: select_image(),
    relief="flat"
)
button_6.place(
    x=31.000000000000114,
    y=255.9999999999999,
    width=90.0,
    height=24.0
)

# ***********************点击进行发送*****************************

def sendmsg(w):
    # 循环发送消息
    if excel_path != "":
        for i, item in enumerate(data):
            notice("第{}个：正在给{}发送".format(i + 1, str(item[1])))
            # 发送文本消息
            w.send_text(self_wx=self_wx, to_wx=str(item[1]), msg=str(item[2]))
            notice("文本信息发送成功")
            time.sleep(1)

            if image_path != "":
                # 发送图片消息
                w.send_img(self_wx=self_wx, to_wx=str(item[1]), path=image_path)
                time.sleep(1)
                notice("图片发送成功")
        notice("----------------------")
    else:
        notice("请选择您要发送的信息或图片")

# 确认发送按钮 button_4
button_image_4 = PhotoImage(
    file=relative_to_assets("button_4.png"))
button_4 = Button(
    image=button_image_4,
    borderwidth=0,
    highlightthickness=0,
    # command=lambda: notice("button_4 clicked"),
    command=lambda: thread_it(sendmsg,w),
    relief="flat"
)
button_4.place(
    x=147.0000000000001,
    y=299.9999999999999,
    width=120.0,
    height=33.0
)

# ***********************生成好友头像照片墙*****************************

friendlists_path = os.path.join(BASE_DIR, "friends_list.xlsx")
avatar_path = os.path.join(BASE_DIR, "images")

# 下载头像图片
def show_download_image():
    notice("正在下载微信好友头像中，请稍候")
    time.sleep(0.5)

def download_image():
    # 如果不存在文件，就先导出,否则读取
    # if not os.path.exists(friendlists_path):
    printFriendBegin()
    friendList(w)

    df = pd.read_excel(friendlists_path)

    if not os.path.exists(avatar_path):
        os.makedirs(avatar_path)
        notice("存放头像图片的images文件夹创建完成")
    else:
        notice("存放头像图片的images文件夹已经存在")

    # 开始计数，循环
    show_download_image()
    count = 0
    for index, row in df.iterrows():
        if row["avatar_url"] != "":
            if not os.path.exists(os.path.join(BASE_DIR, "images", f"{row['wx_id']}.jpeg")):
                try:
                    response = requests.get(row["avatar_url"])
                    open(os.path.join(BASE_DIR, "images", f"{row['wx_id']}.jpeg"), "wb").write(response.content)
                    count += 1
                    notice(f"正在处理第{index+1}个图片： {row['wx_id']}.jpeg已经下载")
                except (requests.exceptions.MissingSchema, requests.exceptions.ConnectionError):
                    notice(f"正在处理第{index+1}个图片： {row['wx_id']}.jpeg链接无效，已跳过")
                    count += 1
                    pass
            else:
                notice(f"正在处理第{index + 1}个图片： {row['wx_id']}.jpeg已经存在")
                count += 1
    notice(f"{count} 个好友头像图片处理完成")
    notice("----------------------")

# 开始进行排列
avatar_square = os.path.join(BASE_DIR, "avatar_square.png")
avatar_love = os.path.join(BASE_DIR, "avatar_love.png")
avatar_character = os.path.join(BASE_DIR, "avatar_character.png")


def show_joint_avatar_square():
    notice("正在拼接方形照片墙中，请稍候")
    time.sleep(0.5)

# 拼接方形头像
def joint_avatar_square():
    # 获取文件夹内头像个数
    length = len(os.listdir(avatar_path))
    notice(f'读取images文件夹成功，共有 {length} 张头像图片')
    show_joint_avatar_square()
    # 拼接后的图片大小
    image_size = 2560
    # 设置每个头像大小
    each_size = math.ceil(image_size / math.floor(math.sqrt(length)))
    # 计算所需各行列的头像数量
    x_lines = math.ceil(math.sqrt(length))
    y_lines = math.ceil(math.sqrt(length))
    image = Image.new('RGB', (each_size * x_lines, each_size * y_lines))
    x = 0
    y = 0
    # 循环拼接
    for image_file in os.listdir(avatar_path):
        try:
            with Image.open(os.path.join(avatar_path, image_file)) as img:
                img = img.resize((each_size, each_size))
                image.paste(img, (x * each_size, y * each_size))
                x += 1
                if x == x_lines:
                    x = 0
                    y += 1
        except IOError:
            notice(f"此头像读取失败：{os.path.join(avatar_path, image_file)}")

    img = image.save(avatar_square)
    notice(f'微信好友头像方形照片墙拼接成功，路径为：{avatar_square} ')
    notice("----------------------")

# 拼接心形照片墙
# 计算心形，判断图像的坐标是否在心形函数内 512**2=262144 1024**2=1048576
def get_heart_shape(x,y):
    y1 = 0.618 * np.abs(x) - 0.7 * np.sqrt(1048576 - x ** 2)
    y2 = 0.618 * np.abs(x) + 0.7 * np.sqrt(1048576 - x ** 2)
    if y<y1 or y>y2:
        return False
    else:
        return True

def show_joint_avatar_love():
    notice("正在拼接心形照片墙中，请稍候")
    time.sleep(0.5)

# 拼接为心形照片墙
def joint_avatar_love():

    joint_avatar_square()

    show_joint_avatar_love()
    # 获取文件夹内头像个数
    length = len(os.listdir(avatar_path))
    # 设定每个头像的大小
    image_size = 2048
    each_size = int(math.sqrt(float(image_size * image_size) / length))
    # 一行的图像个数,若为偶数个，则+1转换为奇数个，修改每个图像的大小，使最后的心形对称好看
    num = int(image_size / each_size)
    if num % 2 == 0:
        num += 1
    each_size = int(image_size / num)
    # 照片墙的行数
    lines = int(image_size / each_size)
    # 创建Image对象，初始化大小,其大小不直接设定为（1024*1024），因为照片拼接出的尺寸不是正好等于1024，故用实际拼接的尺寸
    image = Image.new('RGBA', (lines * each_size, lines * each_size))
    x, y = 0, 0
    for image_file in os.listdir(avatar_path):
        try:
            try:
                # 由于图像坐标从（0，0）开始，而心形函数对应左上角坐标为（-512，512），故此在判断坐标时，稍作转换
                is_heart_part = get_heart_shape((-image_size/2) + x * each_size, (image_size/2) - y * each_size)
                if not is_heart_part:
                    pass
                else:
                    img = Image.open(os.path.join(avatar_path, image_file))
                    # 重新设置图像大小
                    img = img.resize((each_size, each_size), Image.BICUBIC)
                    # 根据x,y坐标位置拼接图像
                    image.paste(img, (x * each_size, y * each_size))
                # 更新下一张图像位置
                x += 1
            except:
                pass
            finally:
                # 一行一行拼接
                if x == lines:
                    x = 0
                    y += 1
        except IOError:
            notice(f"此头像读取失败：{os.path.join(avatar_path, image_file)}")

    img = image.save(avatar_love)
    notice(f'心形照片墙拼接成功，路径为：{avatar_love} ')
    notice("----------------------")


def joint_all():
    download_image()
    joint_avatar_love()

# 按钮 button_3
button_image_3 = PhotoImage(
    file=relative_to_assets("button_3.png"))
button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: thread_it(joint_all),
    relief="flat"
)
button_3.place(
    x=511.0000000000001,
    y=107.99999999999989,
    width=234.0,
    height=33.0
)


# ***********************关注交流，说明文档*****************************
def personalWebsite(event):
    time.sleep(0.5)
    websiteLink = (
        "https://ferryxie.com")
    webbrowser.open_new_tab(websiteLink)

def wechatAccount(event): # 跳转事件
    time.sleep(0.5)
    wechatLink = (
        "https://github.com/ParthJadhav/Tkinter-Designer/"
        "blob/master/docs/instructions.md")
    webbrowser.open_new_tab(wechatLink)

def know_more_clicked(event):
    time.sleep(0.5)
    instructions = ("https://ferryxie.com/archives/4116")
    webbrowser.open_new_tab(instructions)

# button7 公众号
button_image_7 = PhotoImage(
    file=relative_to_assets("button_7.png"))
button_7 = Button(
    image=button_image_7,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: notice("正在跳转去Ferry的个人网站：https://ferryxie.com"),
    relief="flat"
)
button_7.place(
    x=200.0000000000001,
    y=417.9999999999999,
    width=173.0,
    height=28.0
)
button_7.bind('<Button-1>', personalWebsite)

# button8 个人网站
button_image_8 = PhotoImage(
    file=relative_to_assets("button_8.png"))
button_8 = Button(
    image=button_image_8,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: notice("正在跳转去Ferry的个人网站：https://ferryxie.com"),
    relief="flat"
)
button_8.place(
    x=201.0000000000001,
    y=457.9999999999999,
    width=172.0,
    height=28.0
)
button_8.bind('<Button-1>', personalWebsite)

# button_9 说明文档
button_image_9 = PhotoImage(
    file=relative_to_assets("button_9.png"))
button_9 = Button(
    image=button_image_9,
    borderwidth=0,
    highlightthickness=0,
    command=lambda: notice("正在跳转去软件说明文档:https://ferryxie.com/archives/4116"),
    relief="flat"
)
button_9.place(
    x=750.0000000000001,
    y=12.999999999999886,
    width=88.0,
    height=25.0
)
button_9.bind('<Button-1>', know_more_clicked)

# ***********************其他 Tkinter GUI 界面*****************************

canvas.create_rectangle(
    1.1368683772161603e-13,
    1.1368683772161603e-13,
    862.0000000000001,
    46.000000000000114,
    fill="#EDEDED",
    outline="")

canvas.create_rectangle(
    1.1368683772161603e-13,
    44.999999999999886,
    862.0000000000001,
    45.999999999999886,
    fill="#979797",
    outline="")

canvas.create_text(
    394.0000000000001,
    11.999999999999886,
    anchor="nw",
    text="微信好友助手",
    fill="#1E1E1E",
    font=("PingFangSC Medium", 17 * -1)
)

image_image_1 = PhotoImage(
    file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(
    380.0000000000001,
    23.999999999999886,
    image=image_image_1
)

canvas.create_rectangle(
    19.000000000000114,
    60.999999999999886,
    394.0000000000001,
    158.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    31.000000000000114,
    70.99999999999989,
    anchor="nw",
    text="登录｜导出好友列表",
    fill="#1E1E1E",
    font=("microsoft yahei", 12 * -1)
)

canvas.create_rectangle(
    31.000000000000114,
    89.49999999999989,
    390.0000000000001,
    90.49999999999989,
    fill="#979797",
    outline="")

canvas.create_rectangle(
    413.0000000000001,
    60.999999999999886,
    843.0000000000001,
    158.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    435.0000000000001,
    70.99999999999989,
    anchor="nw",
    text="一键生成好友头像照片墙",
    fill="#1E1E1E",
    font=("microsoft yahei", 12 * -1)
)

canvas.create_rectangle(
    435.0000000000001,
    89.49999999999989,
    842.0000000000001,
    90.49999999999989,
    fill="#979797",
    outline="")

canvas.create_rectangle(
    19.000000000000114,
    166.9999999999999,
    394.0000000000001,
    352.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    31.000000000000114,
    176.9999999999999,
    anchor="nw",
    text="发送文本/图片信息",
    fill="#1E1E1E",
    font=("microsoft yahei", 12 * -1)
)

canvas.create_rectangle(
    31.000000000000114,
    195.4999999999999,
    390.0000000000001,
    196.4999999999999,
    fill="#979797",
    outline="")

canvas.create_rectangle(
    19.000000000000114,
    361.9999999999999,
    394.0000000000001,
    508.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    31.000000000000114,
    371.9999999999999,
    anchor="nw",
    text="关注交流｜扫码或点击",
    fill="#1E1E1E",
    font=("microsoft yahei", 12 * -1)
)

canvas.create_rectangle(
    31.000000000000114,
    404.9999999999999,
    378.0000000000001,
    499.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_rectangle(
    31.000000000000114,
    390.4999999999999,
    390.0000000000001,
    391.4999999999999,
    fill="#979797",
    outline="")

image_image_2 = PhotoImage(
    file=relative_to_assets("image_2.png"))
image_2 = canvas.create_image(
    149.0000000000001,
    452.9999999999999,
    image=image_image_2
)

image_image_3 = PhotoImage(
    file=relative_to_assets("image_3.png"))
image_3 = canvas.create_image(
    65.00000000000011,
    451.9999999999999,
    image=image_image_3
)

canvas.create_rectangle(
    413.0000000000001,
    166.9999999999999,
    843.0000000000001,
    508.9999999999999,
    fill="#FFFFFF",
    outline="")

canvas.create_text(
    436.0000000000001,
    176.9999999999999,
    anchor="nw",
    text="运行日志",
    fill="#1E1E1E",
    font=("microsoft yahei", 12 * -1)
)

canvas.create_rectangle(
    436.0000000000001,
    195.9999999999999,
    843.0000000000001,
    196.9999999999999,
    fill="#979797",
    outline="")

canvas.create_rectangle(
    561.0000000000001,
    499.9999999999999,
    695.0000000000001,
    502.9999999999999,
    fill="#000000",
    outline="")

# ***********************主要事件循环*****************************

window.resizable(False, False)
window.mainloop()

# 处理消息回调
# 处理消息回调while True:
#     msg = msg_queue.get()
